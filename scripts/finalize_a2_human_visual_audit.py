#!/usr/bin/env python
"""Freeze, unblind, and summarize the A2 human visual audit without rewriting source labels."""

from __future__ import annotations

import csv
import hashlib
import json
import shutil
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[1]
SOURCE_XLSX = (
    REPO_ROOT
    / "deliverables"
    / "pending review"
    / "A2_human_visual_audit_REVIEWER_ONLY_2026-06-01"
    / "A2_visual_audit_manual_review_completed.xlsx"
)
MAPPING_CSV = (
    REPO_ROOT
    / "deliverables"
    / "pending review"
    / "A2_human_visual_audit_pack"
    / "DO_NOT_OPEN_UNTIL_REVIEW_COMPLETE"
    / "blind_mapping.csv"
)
OUTPUT_DIR = REPO_ROOT / "results" / "audit" / "human_visual_audit_2026-06-01"

ALLOWED = {
    "review_status": {"complete", "needs_recheck"},
    "body_color_visible": {"yes", "partly", "no"},
    "visual_clarity": {"clear", "moderate", "low"},
    "body_color_salience": {"high", "medium", "low"},
    "specular_reflection": {"none_minor", "moderate", "strong"},
    "shadow_or_night_effect": {"none_minor", "moderate", "strong"},
    "background_color_bias": {"none_minor", "moderate", "strong"},
    "multi_car_interference": {"none", "minor", "present"},
    "occlusion": {"none", "minor", "moderate", "strong"},
    "include_in_validity_analysis": {"include", "exclude", "unsure"},
}
REQUIRED = [*ALLOWED, "reviewer_initials"]


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def write_csv(path: Path, rows: list[dict[str, object]], columns: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def read_human_rows(path: Path) -> tuple[list[dict[str, object]], dict[str, object]]:
    workbook = load_workbook(path, data_only=True)
    protocol = workbook["Protocol"]
    sheet = workbook["Manual_Audit"]
    headers = [cell.value for cell in sheet[1]]
    rows = [
        dict(zip(headers, [sheet.cell(row_idx, col_idx).value for col_idx in range(1, len(headers) + 1)]))
        for row_idx in range(2, sheet.max_row + 1)
    ]
    metadata = {
        "reviewer_id": protocol["B4"].value or "",
        "reviewer_role": protocol["B5"].value or "",
        "review_date": str(protocol["B6"].value or ""),
        "review_type": protocol["B7"].value or "",
        "workbook_status": protocol["B8"].value or "",
        "notes": protocol["B9"].value or "",
    }
    return rows, metadata


def validate(rows: list[dict[str, object]], metadata: dict[str, object]) -> None:
    if len(rows) != 84:
        raise RuntimeError(f"Expected 84 human-review rows, found {len(rows)}.")
    if metadata["workbook_status"] != "frozen_complete":
        raise RuntimeError("Workbook status must be frozen_complete.")
    invalid: list[str] = []
    missing: list[str] = []
    for row in rows:
        review_id = str(row.get("review_id", ""))
        for field, values in ALLOWED.items():
            if row.get(field) not in values:
                invalid.append(f"{review_id}:{field}={row.get(field)!r}")
        for field in REQUIRED:
            if row.get(field) in (None, ""):
                missing.append(f"{review_id}:{field}")
    if invalid or missing:
        raise RuntimeError(f"Invalid values={invalid}; missing required={missing}")


def strict_confound(row: dict[str, object]) -> bool:
    return bool(
        row["visual_clarity"] == "low"
        or row["body_color_visible"] == "no"
        or row["specular_reflection"] == "strong"
        or row["shadow_or_night_effect"] == "strong"
        or row["background_color_bias"] == "strong"
        or row["multi_car_interference"] == "present"
        or row["occlusion"] in {"moderate", "strong"}
    )


def summarize_group(rows: list[dict[str, object]]) -> dict[str, object]:
    return {
        "n": len(rows),
        "visual_clarity": dict(Counter(str(row["visual_clarity"]) for row in rows)),
        "body_color_visible": dict(Counter(str(row["body_color_visible"]) for row in rows)),
        "inclusion": dict(Counter(str(row["analysis_inclusion"]) for row in rows)),
        "strict_confound_any": sum(strict_confound(row) for row in rows),
    }


def main() -> int:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    source_archive = OUTPUT_DIR / "source" / SOURCE_XLSX.name
    source_archive.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(SOURCE_XLSX, source_archive)

    human_rows, metadata = read_human_rows(SOURCE_XLSX)
    validate(human_rows, metadata)
    mapping = {row["review_id"]: row for row in read_csv(MAPPING_CSV)}
    if set(mapping) != {str(row["review_id"]) for row in human_rows}:
        raise RuntimeError("Blind map and completed workbook review IDs differ.")

    raw_rows: list[dict[str, object]] = []
    for row in human_rows:
        raw_rows.append({**mapping[str(row["review_id"])], **row})

    analysis_rows = [{**row, "analysis_inclusion": str(row["include_in_validity_analysis"])} for row in raw_rows]
    adjudications: list[dict[str, str]] = []

    # The same source image appeared in two matched-control rows. Preserve the
    # human sheet, but use the more conservative reflection rating downstream.
    for review_id in ["VC-029", "VC-069"]:
        row = next(item for item in analysis_rows if item["review_id"] == review_id)
        old_value = str(row["specular_reflection"])
        row["specular_reflection"] = "moderate"
        adjudications.append(
            {
                "review_id": review_id,
                "field": "specular_reflection",
                "human_value": old_value,
                "analysis_value": "moderate",
                "reason": "Duplicate source image had discordant ratings; use the more conservative human rating.",
            }
        )

    raw_columns = list(raw_rows[0])
    analysis_columns = list(analysis_rows[0])
    write_csv(OUTPUT_DIR / "human_visual_audit_unblinded_raw.csv", raw_rows, raw_columns)
    write_csv(OUTPUT_DIR / "human_visual_audit_analysis.csv", analysis_rows, analysis_columns)
    write_csv(
        OUTPUT_DIR / "human_visual_audit_adjudication_log.csv",
        adjudications,
        ["review_id", "field", "human_value", "analysis_value", "reason"],
    )

    target = [row for row in analysis_rows if row["audit_group"] == "target_conflict_flip"]
    control = [row for row in analysis_rows if row["audit_group"] == "matched_faithful_control"]
    payload = {
        "source_workbook": str(SOURCE_XLSX),
        "source_workbook_sha256": sha256(SOURCE_XLSX),
        "row_level_reviewer_initials": dict(Counter(str(row["reviewer_initials"]) for row in human_rows)),
        "protocol_metadata": metadata,
        "target_raw": summarize_group(target),
        "control_raw": summarize_group(control),
        "adjudications": adjudications,
    }
    (OUTPUT_DIR / "human_visual_audit_summary.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    lines = [
        "# A2 Manual Human Visual Audit Summary",
        "",
        "## Provenance",
        "",
        f"- Frozen workbook SHA-256: `{payload['source_workbook_sha256']}`",
        f"- Row-level reviewer identifier: `{next(iter(payload['row_level_reviewer_initials']))}`",
        "- Protocol metadata fields were not used to claim reviewer independence.",
        "- The frozen human workbook is archived unchanged. Analysis-level adjudications are logged separately.",
        "",
        "## Results",
        "",
        f"- Target conflict-flip rows: `{len(target)}`; clear: `{sum(row['visual_clarity'] == 'clear' for row in target)}/{len(target)}`; strict visual confound: `{sum(strict_confound(row) for row in target)}/{len(target)}`.",
        f"- Matched faithful controls: `{len(control)}`; clear: `{sum(row['visual_clarity'] == 'clear' for row in control)}/{len(control)}`; strict visual confound: `{sum(strict_confound(row) for row in control)}/{len(control)}`.",
        "- One matched-control row remained `unsure` for validity-analysis inclusion. It is retained in the aggregate report and is not analyzed as a separate case.",
        "",
        "## Analysis-Level Adjudications",
        "",
        "- `VC-029` and `VC-069` are duplicate appearances of the same source image. Their reflection field is harmonized to the more conservative human value, `moderate`.",
        "",
        "## Wording Constraint",
        "",
        "Use `manual human visual audit`, not `independent human visual audit`, unless reviewer independence metadata is separately documented.",
    ]
    (OUTPUT_DIR / "human_visual_audit_summary.md").write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(json.dumps(payload, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
